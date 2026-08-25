import os
import json
import unicodedata
import uuid
from io import BytesIO
from contextlib import asynccontextmanager

import cv2
from fastapi import Depends, FastAPI, File, Form, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from sqlalchemy import text
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy.orm import Session

import models
from database import engine, get_db
from processador_imagem import ler_respostas_grade_fixa, processar_folha


UPLOAD_DIR = os.getenv("UPLOAD_DIR", "/tmp/uploads" if os.getenv("VERCEL") else "uploads")
ALTERNATIVAS_VALIDAS = {"A", "B", "C", "D", "X"}
GABARITO_PADRAO = "PADRAO"
GABARITO_ADAPTADA = "ADAPTADA"
ORDEM_DISCIPLINAS_RESULTADO = [
    "PORTUGUES",
    "HISTORIA",
    "GEOGRAFIA",
    "EDUCACAO FISICA",
    "MATEMATICA",
    "CIENCIAS",
    "ARTES",
    "INGLES",
]
TRANSFERIDOS_RELATORIO_POR_TURMA = {
    "6A": {
        1: "Transferidos/Remanejados",
        5: "Transferidos/Remanejados",
        14: "Transferidos/Remanejados",
        "nomes": {
            "ALICE MARTINS": "Transferidos/Remanejados",
            "ALICE MARTINS SILVA": "Transferidos/Remanejados",
            "CAROLINA AKEMI": "Transferidos/Remanejados",
            "CAROLINA AKEMI ISHIKAWA": "Transferidos/Remanejados",
            "ISABEL SIMOES": "Transferidos/Remanejados",
            "ISABEL SIMOES HOISEL": "Transferidos/Remanejados",
        },
    },
    "6B": {
        "nomes": {
            "LIVIA KEIKO": "Transferidos/Remanejados",
            "LIVIA KEIKO RIBEIRO HIRAO": "Transferidos/Remanejados",
        },
    },
    "7B": {
        3: "Transferidos/Remanejados",
        11: "Transferidos/Remanejados",
        22: "Transferidos/Remanejados",
        "nomes": {
            "ARTHUR JESUS": "Transferidos/Remanejados",
            "ARTHUR JESUS DA SILVA": "Transferidos/Remanejados",
            "GIULIA PAPIN": "Transferidos/Remanejados",
            "GIULIA PAPIN SILVA": "Transferidos/Remanejados",
            "MILLENA CAMPOS": "Transferidos/Remanejados",
            "MILLENA CAMPOS DE AZEVEDO SILVA": "Transferidos/Remanejados",
        },
    },
    "8A": {
        13: "Transferido",
        17: "Transferido",
        19: "Transferido",
    },
    "8B": {
        10: "Transferida/Remanejada 8B",
    },
    "8C": {
        4: "Transferidos/Remanejado",
        18: "Transferidos/Remanejado",
    },
    "9A": {
        13: "Remanejados/Transferidos",
    },
    "9C": {
        5: "Transferidos/Remanejados",
        "nomes": {
            "REBECA SANTOS": "Transferidos/Remanejados",
            "REBECA SANTOS BASILIO": "Transferidos/Remanejados",
        },
    },
}

os.makedirs(UPLOAD_DIR, exist_ok=True)


def _inicializar_banco():
    if engine is None:
        print("DATABASE_URL nao configurada; inicializacao do banco ignorada.")
        return

    try:
        models.Base.metadata.create_all(bind=engine)
        _preparar_banco()
    except SQLAlchemyError as exc:
        print(f"Erro ao inicializar banco: {exc}")


@asynccontextmanager
async def lifespan(app: FastAPI):
    _inicializar_banco()
    yield


app = FastAPI(title="Sistema de Correcao de Gabaritos", lifespan=lifespan)

app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:5173",
        "http://127.0.0.1:5173",
        "https://corretor-front-kappa.vercel.app",
    ],
    allow_origin_regex=r"https://.*\.vercel\.app",
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


def _preparar_banco():
    with engine.begin() as conn:
        conn.execute(
            text(
                """
                alter table if exists gabaritos
                drop constraint if exists gabaritos_modelo_prova_id_numero_questao_key
                """
            )
        )
        conn.execute(
            text(
                """
                alter table if exists gabaritos
                drop constraint if exists uq_gabaritos_modelo_serie_questao
                """
            )
        )
        conn.execute(
            text(
                """
                alter table if exists gabaritos
                add column if not exists codigo_gabarito varchar not null default 'PADRAO'
                """
            )
        )
        conn.execute(
            text(
                """
                alter table if exists resultados_alunos
                add column if not exists codigo_gabarito varchar not null default 'PADRAO'
                """
            )
        )
        conn.execute(
            text(
                """
                alter table if exists gabaritos
                drop constraint if exists gabaritos_resposta_correta_check
                """
            )
        )
        conn.execute(
            text(
                """
                alter table if exists gabaritos
                add constraint gabaritos_resposta_correta_check
                check (resposta_correta in ('A', 'B', 'C', 'D', 'X'))
                """
            )
        )
        conn.execute(
            text(
                """
                alter table if exists respostas_alunos
                drop constraint if exists respostas_alunos_resposta_aluno_check
                """
            )
        )
        conn.execute(
            text(
                """
                alter table if exists respostas_alunos
                add constraint respostas_alunos_resposta_aluno_check
                check (resposta_aluno in ('A', 'B', 'C', 'D', 'X'))
                """
            )
        )
        conn.execute(
            text(
                """
                alter table if exists respostas_alunos
                drop constraint if exists respostas_alunos_resposta_correta_check
                """
            )
        )
        conn.execute(
            text(
                """
                alter table if exists respostas_alunos
                add constraint respostas_alunos_resposta_correta_check
                check (resposta_correta in ('A', 'B', 'C', 'D', 'X'))
                """
            )
        )
        conn.execute(
            text(
                """
                do $$
                begin
                    if not exists (
                        select 1
                        from pg_constraint
                        where conname = 'uq_gabaritos_modelo_serie_codigo_questao'
                    ) then
                        alter table gabaritos
                        add constraint uq_gabaritos_modelo_serie_codigo_questao
                        unique (modelo_prova_id, serie, codigo_gabarito, numero_questao);
                    end if;
                end $$;
                """
            )
        )


if engine is not None:
    _preparar_banco()


def _buscar_modelo_prova(db: Session, escola_id: str, bimestre: int, dia: int):
    modelo = (
        db.query(models.ModeloProva)
        .filter(models.ModeloProva.escola_id == escola_id)
        .filter(models.ModeloProva.bimestre == bimestre)
        .filter(models.ModeloProva.dia == dia)
        .first()
    )

    if not modelo:
        raise HTTPException(status_code=404, detail="Modelo de prova nao encontrado")

    return modelo


def _extrair_serie_turma(nome_turma: str):
    if not nome_turma:
        return None

    digitos = "".join(caractere for caractere in nome_turma if caractere.isdigit())
    return int(digitos) if digitos else None


def _normalizar_texto(valor: str):
    texto = unicodedata.normalize("NFKD", valor or "")
    texto = "".join(caractere for caractere in texto if not unicodedata.combining(caractere))
    return texto.upper()


def _status_transferencia_relatorio(
    nome_turma: str,
    numero_chamada: int | None,
    nome_aluno: str = "",
):
    turma = _normalizar_texto(nome_turma).replace(" ", "")
    transferidos_turma = TRANSFERIDOS_RELATORIO_POR_TURMA.get(turma, {})
    status_por_numero = transferidos_turma.get(numero_chamada)
    if status_por_numero:
        return status_por_numero

    nomes_transferidos = transferidos_turma.get("nomes", {})
    aluno_normalizado = _normalizar_texto(nome_aluno)
    return nomes_transferidos.get(aluno_normalizado)


def _normalizar_disciplina_ordem(valor: str):
    texto = _normalizar_texto(valor).replace(".", " ")
    texto = " ".join(texto.split())

    equivalencias = {
        "LP": "PORTUGUES",
        "LINGUA PORTUGUESA": "PORTUGUES",
        "PORTUGUES": "PORTUGUES",
        "HIS": "HISTORIA",
        "HISTORIA": "HISTORIA",
        "GEO": "GEOGRAFIA",
        "GEOGRAFIA": "GEOGRAFIA",
        "EF": "EDUCACAO FISICA",
        "ED FISICA": "EDUCACAO FISICA",
        "EDUCACAO FISICA": "EDUCACAO FISICA",
        "MAT": "MATEMATICA",
        "MATEMATICA": "MATEMATICA",
        "CIE": "CIENCIAS",
        "CIENCIAS": "CIENCIAS",
        "ARTES": "ARTES",
        "INGLES": "INGLES",
    }

    return equivalencias.get(texto, texto)


def _ordenar_disciplinas_resultado(disciplinas):
    def chave(disciplina):
        disciplina_normalizada = _normalizar_disciplina_ordem(disciplina)
        try:
            return (ORDEM_DISCIPLINAS_RESULTADO.index(disciplina_normalizada), disciplina)
        except ValueError:
            return (len(ORDEM_DISCIPLINAS_RESULTADO), disciplina)

    return sorted(disciplinas, key=chave)


def _montar_status_resultado_final(escola_nome: str, dias_modelo, dias_corrigidos):
    if not dias_corrigidos:
        return "Pendente"

    dias_modelo = sorted({int(dia) for dia in dias_modelo if dia is not None})
    dias_corrigidos = {int(dia) for dia in dias_corrigidos if dia is not None}
    dias_faltantes = [dia for dia in dias_modelo if dia not in dias_corrigidos]

    if len(dias_modelo) > 1 and len(dias_faltantes) == 1:
        return f"Falta Dia {dias_faltantes[0]}"

    if len(dias_modelo) > 1 and len(dias_faltantes) > 1:
        return "Faltam Dias " + ", ".join(str(dia) for dia in dias_faltantes)

    return "Corrigido"


def _codigo_gabarito_turma(nome_escola: str, nome_turma: str, serie: int, bimestre: int, dia: int):
    escola_normalizada = _normalizar_texto(nome_escola)
    turma_normalizada = _normalizar_texto(nome_turma).replace(" ", "")

    if "AGENOR" in escola_normalizada and bimestre == 2:
        if dia == 1:
            return "CADERNO_A"

        if dia == 2:
            return "CADERNO_B"

    if "TAKAOKA" in escola_normalizada and serie == 8 and dia == 1:
        if "8B" in turma_normalizada:
            return "8B"

        if "8A" in turma_normalizada or "8C" in turma_normalizada:
            return "8AC"

    return GABARITO_PADRAO


def _normalizar_codigo_gabarito(codigo: str | None):
    codigo = (codigo or GABARITO_PADRAO).strip().upper()
    return codigo or GABARITO_PADRAO


def _buscar_contexto_aluno(db: Session, aluno_id: str, modelo):
    aluno = db.query(models.Aluno).filter(models.Aluno.id == aluno_id).first()

    if not aluno:
        raise HTTPException(status_code=404, detail="Aluno nao encontrado")

    turma = db.query(models.Turma).filter(models.Turma.id == aluno.turma_id).first()
    serie = _extrair_serie_turma(turma.nome if turma else "")

    escola = db.query(models.Escola).filter(models.Escola.id == modelo.escola_id).first()

    if not serie and "DANIELA" in _normalizar_texto(escola.nome if escola else ""):
        serie = 8

    if not serie:
        raise HTTPException(status_code=400, detail="Nao consegui identificar a serie do aluno")

    codigo_gabarito = _codigo_gabarito_turma(
        escola.nome if escola else "",
        turma.nome if turma else "",
        serie,
        modelo.bimestre,
        modelo.dia,
    )

    return serie, codigo_gabarito


def _buscar_contexto_correcao(db: Session, aluno_id: str, modelo, codigo_gabarito: str | None = None):
    serie, codigo_padrao = _buscar_contexto_aluno(db, aluno_id, modelo)
    codigo_final = _normalizar_codigo_gabarito(codigo_gabarito) if codigo_gabarito else codigo_padrao
    return serie, codigo_final


def _modelo_eh_daniela(db: Session, modelo):
    escola = db.query(models.Escola).filter(models.Escola.id == modelo.escola_id).first()
    return "DANIELA" in _normalizar_texto(escola.nome if escola else "")


def _buscar_gabarito(db: Session, modelo_id, serie: int, codigo_gabarito: str = GABARITO_PADRAO):
    codigo_gabarito = _normalizar_codigo_gabarito(codigo_gabarito)
    gabaritos = (
        db.query(models.Gabarito)
        .filter(models.Gabarito.modelo_prova_id == modelo_id)
        .filter(models.Gabarito.serie == serie)
        .filter(models.Gabarito.codigo_gabarito == codigo_gabarito)
        .order_by(models.Gabarito.numero_questao)
        .all()
    )

    if not gabaritos:
        raise HTTPException(
            status_code=404,
            detail=f"Gabarito oficial da serie {serie} ({codigo_gabarito}) nao encontrado",
        )

    return gabaritos


def _normalizar_resposta(resposta):
    if resposta is None:
        return None

    resposta = str(resposta).strip().upper()
    return resposta if resposta in ALTERNATIVAS_VALIDAS else None


def _questao_anulada(resposta_correta):
    return _normalizar_resposta(resposta_correta) == "X"


def _salvar_upload(foto: UploadFile, conteudo: bytes):
    if not foto.filename:
        raise HTTPException(status_code=400, detail="Arquivo sem nome")

    extensao = os.path.splitext(foto.filename)[1].lower().replace(".", "")
    if extensao not in {"jpg", "jpeg", "png", "webp", "bmp"}:
        raise HTTPException(
            status_code=400,
            detail="Formato de imagem invalido. Use jpg, jpeg, png, webp ou bmp.",
        )

    if not conteudo:
        raise HTTPException(status_code=400, detail="Arquivo de imagem vazio")

    nome_arquivo = f"{uuid.uuid4()}.{extensao}"
    caminho_arquivo = os.path.join(UPLOAD_DIR, nome_arquivo)

    with open(caminho_arquivo, "wb") as arquivo:
        arquivo.write(conteudo)

    return nome_arquivo, caminho_arquivo


def _comparar_e_salvar_respostas(db: Session, aluno_id: str, modelo_id, gabaritos, respostas_detectadas):
    acertos = 0
    respostas_salvas = []

    (
        db.query(models.RespostaAluno)
        .filter(models.RespostaAluno.aluno_id == aluno_id)
        .filter(models.RespostaAluno.modelo_prova_id == modelo_id)
        .delete(synchronize_session=False)
    )

    for gab in gabaritos:
        resposta_aluno = _normalizar_resposta(
            respostas_detectadas.get(gab.numero_questao)
            or respostas_detectadas.get(str(gab.numero_questao))
        )
        resposta_correta = _normalizar_resposta(gab.resposta_correta)
        acertou = True if _questao_anulada(resposta_correta) else resposta_aluno == resposta_correta

        if acertou:
            acertos += 1

        nova_resposta = models.RespostaAluno(
            aluno_id=aluno_id,
            modelo_prova_id=modelo_id,
            numero_questao=gab.numero_questao,
            disciplina=gab.disciplina,
            resposta_aluno=resposta_aluno,
            resposta_correta=resposta_correta,
            acertou=acertou,
        )

        db.add(nova_resposta)
        respostas_salvas.append(
            {
                "numero_questao": gab.numero_questao,
                "disciplina": gab.disciplina,
                "resposta_aluno": resposta_aluno,
                "resposta_correta": resposta_correta,
                "acertou": acertou,
            }
        )

    return acertos, respostas_salvas


def _calcular_nota(acertos: int, total_questoes: int):
    return round((acertos / total_questoes) * 10, 1) if total_questoes else 0


def _calcular_media_notas_disciplinas(notas_disciplinas):
    notas = [nota for nota in notas_disciplinas.values() if nota is not None]
    return round(sum(notas) / len(notas), 1) if notas else 0


def _calcular_nota_global_bimestre(db: Session, aluno_id: str, escola_id: str, bimestre: int):
    modelos = (
        db.query(models.ModeloProva)
        .filter(models.ModeloProva.escola_id == escola_id)
        .filter(models.ModeloProva.bimestre == bimestre)
        .all()
    )
    modelo_ids = [modelo.id for modelo in modelos]
    resultados = (
        db.query(models.ResultadoAluno)
        .filter(models.ResultadoAluno.aluno_id == aluno_id)
        .filter(models.ResultadoAluno.escola_id == escola_id)
        .filter(models.ResultadoAluno.bimestre == bimestre)
        .all()
    )
    acertos = sum(resultado.acertos for resultado in resultados)
    total_questoes = sum(resultado.total_questoes for resultado in resultados)
    disciplinas = (
        db.query(models.DisciplinaProva)
        .filter(models.DisciplinaProva.modelo_prova_id.in_(modelo_ids))
        .order_by(models.DisciplinaProva.ordem)
        .all()
        if modelo_ids
        else []
    )
    nomes_disciplinas = []
    for disciplina in disciplinas:
        if disciplina.disciplina not in nomes_disciplinas:
            nomes_disciplinas.append(disciplina.disciplina)

    respostas = (
        db.query(models.RespostaAluno)
        .filter(models.RespostaAluno.aluno_id == aluno_id)
        .filter(models.RespostaAluno.modelo_prova_id.in_(modelo_ids))
        .all()
        if modelo_ids
        else []
    )
    resumo_disciplinas = {
        disciplina: {"acertos": 0, "total": 0}
        for disciplina in nomes_disciplinas
    }
    for resposta in respostas:
        disciplina = resposta.disciplina or "Sem disciplina"
        if disciplina not in resumo_disciplinas:
            resumo_disciplinas[disciplina] = {"acertos": 0, "total": 0}
        resumo_disciplinas[disciplina]["total"] += 1
        if resposta.acertou:
            resumo_disciplinas[disciplina]["acertos"] += 1

    notas_disciplinas = {
        disciplina: _calcular_nota(resumo["acertos"], resumo["total"]) if resumo["total"] else 0
        for disciplina, resumo in resumo_disciplinas.items()
    }

    return {
        "acertos_global": acertos,
        "total_questoes_global": total_questoes,
        "nota_global": _calcular_media_notas_disciplinas(notas_disciplinas),
    }


def _buscar_serie_aluno(db: Session, aluno_id: str):
    aluno = db.query(models.Aluno).filter(models.Aluno.id == aluno_id).first()

    if not aluno:
        raise HTTPException(status_code=404, detail="Aluno nao encontrado")

    turma = db.query(models.Turma).filter(models.Turma.id == aluno.turma_id).first()
    serie = _extrair_serie_turma(turma.nome if turma else "")

    if not serie:
        raise HTTPException(status_code=400, detail="Nao consegui identificar a serie do aluno")

    return serie


def _salvar_resultado_aluno(
    db: Session,
    aluno_id: str,
    modelo,
    serie: int,
    codigo_gabarito: str,
    acertos: int,
    total_questoes: int,
):
    nota_dia = _calcular_nota(acertos, total_questoes)

    resultado = (
        db.query(models.ResultadoAluno)
        .filter(models.ResultadoAluno.aluno_id == aluno_id)
        .filter(models.ResultadoAluno.modelo_prova_id == modelo.id)
        .first()
    )

    if not resultado:
        resultado = models.ResultadoAluno(
            aluno_id=aluno_id,
            modelo_prova_id=modelo.id,
            escola_id=modelo.escola_id,
            bimestre=modelo.bimestre,
            dia=modelo.dia,
            serie=serie,
            codigo_gabarito=codigo_gabarito,
            acertos=acertos,
            total_questoes=total_questoes,
            nota_global=nota_dia,
        )
        db.add(resultado)
    else:
        resultado.escola_id = modelo.escola_id
        resultado.bimestre = modelo.bimestre
        resultado.dia = modelo.dia
        resultado.serie = serie
        resultado.codigo_gabarito = codigo_gabarito
        resultado.acertos = acertos
        resultado.total_questoes = total_questoes
        resultado.nota_global = nota_dia

    db.flush()

    return nota_dia


def _recalcular_resultados_por_gabarito(db: Session, modelo, serie: int, codigo_gabarito: str, gabaritos):
    resultados = (
        db.query(models.ResultadoAluno)
        .filter(models.ResultadoAluno.modelo_prova_id == modelo.id)
        .filter(models.ResultadoAluno.serie == serie)
        .filter(models.ResultadoAluno.codigo_gabarito == codigo_gabarito)
        .all()
    )

    gabaritos_por_numero = {gabarito.numero_questao: gabarito for gabarito in gabaritos}
    recalculados = 0

    for resultado in resultados:
        respostas = (
            db.query(models.RespostaAluno)
            .filter(models.RespostaAluno.aluno_id == resultado.aluno_id)
            .filter(models.RespostaAluno.modelo_prova_id == modelo.id)
            .order_by(models.RespostaAluno.numero_questao)
            .all()
        )

        if not respostas:
            continue

        respostas_por_numero = {resposta.numero_questao: resposta for resposta in respostas}
        acertos = 0

        for numero_questao, gabarito in gabaritos_por_numero.items():
            resposta = respostas_por_numero.get(numero_questao)
            resposta_correta = _normalizar_resposta(gabarito.resposta_correta)

            if not resposta:
                resposta = models.RespostaAluno(
                    aluno_id=resultado.aluno_id,
                    modelo_prova_id=modelo.id,
                    numero_questao=numero_questao,
                    resposta_aluno=None,
                )
                db.add(resposta)

            resposta.disciplina = gabarito.disciplina
            resposta.resposta_correta = resposta_correta
            resposta.acertou = (
                True
                if _questao_anulada(resposta_correta)
                else _normalizar_resposta(resposta.resposta_aluno) == resposta_correta
            )

            if resposta.acertou:
                acertos += 1

        resultado.escola_id = modelo.escola_id
        resultado.bimestre = modelo.bimestre
        resultado.dia = modelo.dia
        resultado.serie = serie
        resultado.codigo_gabarito = codigo_gabarito
        resultado.acertos = acertos
        resultado.total_questoes = len(gabaritos_por_numero)
        resultado.nota_global = _calcular_nota(acertos, resultado.total_questoes)
        recalculados += 1

    return recalculados


def _tabular_respostas(respostas):
    respostas_ordenadas = sorted(
        respostas,
        key=lambda resposta: resposta["numero_questao"],
    )

    por_questao = {
        f"q{resposta['numero_questao']}": resposta["resposta_aluno"]
        for resposta in respostas_ordenadas
    }

    por_disciplina = {}

    for resposta in respostas_ordenadas:
        disciplina = resposta["disciplina"] or "Sem disciplina"

        if disciplina not in por_disciplina:
            por_disciplina[disciplina] = {
                "respostas": {},
                "acertos": 0,
                "total": 0,
            }

        por_disciplina[disciplina]["respostas"][f"q{resposta['numero_questao']}"] = resposta[
            "resposta_aluno"
        ]
        por_disciplina[disciplina]["total"] += 1

        if resposta["acertou"]:
            por_disciplina[disciplina]["acertos"] += 1

    for resumo in por_disciplina.values():
        resumo["nota"] = _calcular_nota(resumo["acertos"], resumo["total"])

    return {
        "por_questao": por_questao,
        "por_disciplina": por_disciplina,
        "linhas": respostas_ordenadas,
    }


def _resposta_aluno_para_dict(resposta):
    return {
        "numero_questao": resposta.numero_questao,
        "disciplina": resposta.disciplina,
        "resposta_aluno": resposta.resposta_aluno,
        "resposta_correta": resposta.resposta_correta,
        "acertou": resposta.acertou,
    }


def _valor_planilha(valor):
    return "" if valor is None else valor


def _montar_excel_resultado_final(escola_nome: str, bimestre: int, disciplinas, linhas):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    planilha = workbook.active
    planilha.title = "Resultado final"

    cabecalho = [
        "Turma",
        "Nº",
        "Aluno",
        *disciplinas,
        "Media Global",
        "Status",
    ]
    planilha.append([f"Resultado final - {escola_nome} - {bimestre}º bimestre"])
    planilha.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cabecalho))
    planilha.append(cabecalho)

    for linha in linhas:
        planilha.append(
            [
                linha["turma"],
                linha["numero_chamada"],
                linha["aluno"],
                *[_valor_planilha(linha["disciplinas"].get(disciplina)) for disciplina in disciplinas],
                _valor_planilha(linha["nota_global"]),
                linha["status"],
            ]
        )

    titulo = planilha[1][0]
    titulo.font = Font(bold=True, size=14)
    titulo.alignment = Alignment(horizontal="center")

    preenchimento_cabecalho = PatternFill("solid", fgColor="D9EAF7")
    preenchimento_transferido = PatternFill("solid", fgColor="DCFCE7")
    for celula in planilha[2]:
        celula.font = Font(bold=True)
        celula.fill = preenchimento_cabecalho
        celula.alignment = Alignment(horizontal="center")

    for indice, linha in enumerate(linhas, start=3):
        if not linha.get("transferido"):
            continue

        for celula in planilha[indice]:
            celula.fill = preenchimento_transferido

    for coluna in planilha.columns:
        largura = max(len(str(celula.value or "")) for celula in coluna) + 2
        planilha.column_dimensions[get_column_letter(coluna[0].column)].width = min(max(largura, 12), 34)

    for linha in planilha.iter_rows(min_row=3):
        for celula in linha:
            celula.alignment = Alignment(vertical="center")

    arquivo = BytesIO()
    workbook.save(arquivo)
    arquivo.seek(0)
    return arquivo


def _montar_resultado_final_escola(db: Session, escola_id: str, bimestre: int):
    escola = db.query(models.Escola).filter(models.Escola.id == escola_id).first()
    escola_nome = escola.nome if escola else ""
    modelos = (
        db.query(models.ModeloProva)
        .filter(models.ModeloProva.escola_id == escola_id)
        .filter(models.ModeloProva.bimestre == bimestre)
        .order_by(models.ModeloProva.dia)
        .all()
    )

    if not modelos:
        raise HTTPException(status_code=404, detail="Modelo de prova nao encontrado")

    modelo_ids = [modelo.id for modelo in modelos]
    dias_modelo = [modelo.dia for modelo in modelos]
    turmas = (
        db.query(models.Turma)
        .filter(models.Turma.escola_id == escola_id)
        .order_by(models.Turma.nome)
        .all()
    )
    turma_ids = [turma.id for turma in turmas]
    turma_por_id = {turma.id: turma for turma in turmas}

    alunos = (
        db.query(models.Aluno)
        .filter(models.Aluno.turma_id.in_(turma_ids))
        .order_by(models.Aluno.numero_chamada)
        .all()
        if turma_ids
        else []
    )
    aluno_ids = [aluno.id for aluno in alunos]

    respostas = (
        db.query(models.RespostaAluno)
        .filter(models.RespostaAluno.modelo_prova_id.in_(modelo_ids))
        .filter(models.RespostaAluno.aluno_id.in_(aluno_ids))
        .all()
        if aluno_ids
        else []
    )
    resultados = (
        db.query(models.ResultadoAluno)
        .filter(models.ResultadoAluno.modelo_prova_id.in_(modelo_ids))
        .filter(models.ResultadoAluno.aluno_id.in_(aluno_ids))
        .all()
        if aluno_ids
        else []
    )

    disciplinas_modelo = (
        db.query(models.DisciplinaProva)
        .filter(models.DisciplinaProva.modelo_prova_id.in_(modelo_ids))
        .order_by(models.DisciplinaProva.ordem)
        .all()
        if modelo_ids
        else []
    )
    disciplinas = []
    disciplinas_por_modelo = {}
    for disciplina in disciplinas_modelo:
        disciplinas_por_modelo.setdefault(disciplina.modelo_prova_id, []).append(
            disciplina.disciplina
        )
        if disciplina.disciplina not in disciplinas:
            disciplinas.append(disciplina.disciplina)

    resumo_disciplinas = {}
    for resposta in respostas:
        aluno_id = str(resposta.aluno_id)
        disciplina = resposta.disciplina or "Sem disciplina"

        if disciplina not in disciplinas:
            disciplinas.append(disciplina)

        if aluno_id not in resumo_disciplinas:
            resumo_disciplinas[aluno_id] = {}

        if disciplina not in resumo_disciplinas[aluno_id]:
            resumo_disciplinas[aluno_id][disciplina] = {"acertos": 0, "total": 0}

        resumo_disciplinas[aluno_id][disciplina]["total"] += 1
        if resposta.acertou:
            resumo_disciplinas[aluno_id][disciplina]["acertos"] += 1

    disciplinas = _ordenar_disciplinas_resultado(disciplinas)

    resumo_global = {}
    dias_corrigidos_por_aluno = {}
    for resultado in resultados:
        aluno_id = str(resultado.aluno_id)

        if aluno_id not in resumo_global:
            resumo_global[aluno_id] = {"acertos": 0, "total": 0}

        resumo_global[aluno_id]["acertos"] += resultado.acertos
        resumo_global[aluno_id]["total"] += resultado.total_questoes
        dias_corrigidos_por_aluno.setdefault(aluno_id, set()).add(resultado.dia)

        if resultado.total_questoes == 0:
            if aluno_id not in resumo_disciplinas:
                resumo_disciplinas[aluno_id] = {}

            for disciplina in disciplinas_por_modelo.get(resultado.modelo_prova_id, []):
                resumo_disciplinas[aluno_id][disciplina] = {
                    "acertos": 0,
                    "total": 0,
                    "nota_manual": resultado.nota_global,
                }

    linhas = []
    for aluno in sorted(
        alunos,
        key=lambda item: (
            turma_por_id.get(item.turma_id).nome if turma_por_id.get(item.turma_id) else "",
            item.numero_chamada,
            item.nome,
        ),
    ):
        aluno_id = str(aluno.id)
        turma = turma_por_id.get(aluno.turma_id)
        global_aluno = resumo_global.get(aluno_id, {"acertos": 0, "total": 0})

        notas_disciplinas = {}
        for disciplina in disciplinas:
            resumo = resumo_disciplinas.get(aluno_id, {}).get(disciplina)
            if resumo and "nota_manual" in resumo:
                notas_disciplinas[disciplina] = resumo["nota_manual"]
            else:
                notas_disciplinas[disciplina] = (
                    _calcular_nota(resumo["acertos"], resumo["total"]) if resumo else 0
                )
        nota_global = _calcular_media_notas_disciplinas(notas_disciplinas)
        status_transferencia = _status_transferencia_relatorio(
            turma.nome if turma else "",
            aluno.numero_chamada,
            aluno.nome,
        )
        transferido = bool(status_transferencia)

        linhas.append(
            {
                "turma": turma.nome if turma else "",
                "numero_chamada": aluno.numero_chamada,
                "aluno": aluno.nome,
                "disciplinas": notas_disciplinas,
                "nota_global": nota_global,
                "transferido": transferido,
                "status": status_transferencia
                if transferido
                else _montar_status_resultado_final(
                    escola_nome,
                    dias_modelo,
                    dias_corrigidos_por_aluno.get(aluno_id, set()),
                ),
            }
        )

    return disciplinas, linhas


def _montar_relatorio_ausentes_turma(db: Session, turma_id: str, escola_id: str, bimestre: int):
    turma = (
        db.query(models.Turma)
        .filter(models.Turma.id == turma_id)
        .filter(models.Turma.escola_id == escola_id)
        .first()
    )
    if not turma:
        raise HTTPException(status_code=404, detail="Turma nao encontrada")

    escola = db.query(models.Escola).filter(models.Escola.id == escola_id).first()
    if not escola:
        raise HTTPException(status_code=404, detail="Escola nao encontrada")

    modelos = (
        db.query(models.ModeloProva)
        .filter(models.ModeloProva.escola_id == escola_id)
        .filter(models.ModeloProva.bimestre == bimestre)
        .order_by(models.ModeloProva.dia)
        .all()
    )
    if not modelos:
        raise HTTPException(status_code=404, detail="Modelo de prova nao encontrado")

    modelo_ids = [modelo.id for modelo in modelos]
    dias_modelo = sorted({modelo.dia for modelo in modelos})

    alunos = (
        db.query(models.Aluno)
        .filter(models.Aluno.turma_id == turma_id)
        .order_by(models.Aluno.numero_chamada, models.Aluno.nome)
        .all()
    )
    aluno_ids = [aluno.id for aluno in alunos]

    resultados = (
        db.query(models.ResultadoAluno)
        .filter(models.ResultadoAluno.modelo_prova_id.in_(modelo_ids))
        .filter(models.ResultadoAluno.aluno_id.in_(aluno_ids))
        .all()
        if aluno_ids
        else []
    )

    dias_realizados_por_aluno = {}
    for resultado in resultados:
        dias_realizados_por_aluno.setdefault(str(resultado.aluno_id), set()).add(resultado.dia)

    linhas = []
    for aluno in alunos:
        aluno_id = str(aluno.id)
        dias_realizados = dias_realizados_por_aluno.get(aluno_id, set())
        dias_ausentes = [dia for dia in dias_modelo if dia not in dias_realizados]

        if not dias_ausentes:
            continue

        status_transferencia = _status_transferencia_relatorio(
            turma.nome,
            aluno.numero_chamada,
            aluno.nome,
        )
        transferido = bool(status_transferencia)
        if transferido:
            continue

        linhas.append(
            {
                "aluno_id": aluno_id,
                "numero_chamada": aluno.numero_chamada,
                "aluno": aluno.nome,
                "transferido": transferido,
                "ausente_dia_1": 1 in dias_ausentes if 1 in dias_modelo else None,
                "ausente_dia_2": 2 in dias_ausentes if 2 in dias_modelo else None,
                "dias_ausentes": dias_ausentes,
                "dias_realizados": sorted(dias_realizados),
                "status": status_transferencia
                if transferido
                else "Faltou " + " e ".join(f"Dia {dia}" for dia in dias_ausentes),
            }
        )

    return {
        "escola": escola.nome,
        "turma": turma.nome,
        "turma_id": turma_id,
        "escola_id": escola_id,
        "bimestre": bimestre,
        "dias_modelo": dias_modelo,
        "total_alunos": len(alunos),
        "total_ausentes": len(linhas),
        "total_presentes": len(alunos) - len(linhas),
        "total_transferidos": sum(1 for linha in linhas if linha["transferido"]),
        "alunos": linhas,
    }


def _montar_excel_relatorio_ausentes(dados):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    planilha = workbook.active
    planilha.title = "Ausentes por turma"

    cabecalho = ["NÂº", "Aluno", "Dia 1", "Dia 2", "Status"]
    planilha.append(
        [
            f"Alunos sem avaliacao - {dados['escola']} - {dados['turma']} - "
            f"{dados['bimestre']}Âº bimestre"
        ]
    )
    planilha.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cabecalho))
    planilha.append(cabecalho)

    for aluno in dados["alunos"]:
        planilha.append(
            [
                aluno["numero_chamada"],
                aluno["aluno"],
                "Nao realizou" if aluno["ausente_dia_1"] else "-" if aluno["ausente_dia_1"] is None else "Realizou",
                "Nao realizou" if aluno["ausente_dia_2"] else "-" if aluno["ausente_dia_2"] is None else "Realizou",
                aluno["status"],
            ]
        )

    titulo = planilha[1][0]
    titulo.font = Font(bold=True, size=14)
    titulo.alignment = Alignment(horizontal="center")

    preenchimento_cabecalho = PatternFill("solid", fgColor="FEE2E2")
    preenchimento_transferido = PatternFill("solid", fgColor="DCFCE7")
    for celula in planilha[2]:
        celula.font = Font(bold=True)
        celula.fill = preenchimento_cabecalho
        celula.alignment = Alignment(horizontal="center")

    for indice, aluno in enumerate(dados["alunos"], start=3):
        if not aluno.get("transferido"):
            continue

        for celula in planilha[indice]:
            celula.fill = preenchimento_transferido

    for coluna in planilha.columns:
        largura = max(len(str(celula.value or "")) for celula in coluna) + 2
        planilha.column_dimensions[get_column_letter(coluna[0].column)].width = min(max(largura, 12), 42)

    for linha in planilha.iter_rows(min_row=3):
        for celula in linha:
            celula.alignment = Alignment(vertical="center")

    arquivo = BytesIO()
    workbook.save(arquivo)
    arquivo.seek(0)
    return arquivo


def _nome_aba_excel(nome: str, usados: set[str]):
    caracteres_invalidos = "[]:*?/\\"
    nome_limpo = "".join("_" if caractere in caracteres_invalidos else caractere for caractere in nome)
    nome_base = (nome_limpo or "Turma").strip()[:31]
    nome_aba = nome_base
    indice = 2

    while nome_aba in usados:
        sufixo = f" {indice}"
        nome_aba = f"{nome_base[:31 - len(sufixo)]}{sufixo}"
        indice += 1

    usados.add(nome_aba)
    return nome_aba


def _montar_excel_relatorios_ausentes_escola(db: Session, escola_id: str, bimestre: int):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    escola = db.query(models.Escola).filter(models.Escola.id == escola_id).first()
    if not escola:
        raise HTTPException(status_code=404, detail="Escola nao encontrada")

    turmas = (
        db.query(models.Turma)
        .filter(models.Turma.escola_id == escola_id)
        .order_by(models.Turma.nome)
        .all()
    )
    if not turmas:
        raise HTTPException(status_code=404, detail="Turma nao encontrada")

    workbook = Workbook()
    abas_usadas = set()
    cabecalho = ["NÃ‚Âº", "Aluno", "Dia 1", "Dia 2", "Status"]
    preenchimento_cabecalho = PatternFill("solid", fgColor="FEE2E2")

    for indice, turma in enumerate(turmas):
        planilha = workbook.active if indice == 0 else workbook.create_sheet()
        planilha.title = _nome_aba_excel(turma.nome, abas_usadas)
        dados = _montar_relatorio_ausentes_turma(db, str(turma.id), escola_id, bimestre)

        planilha.append(
            [
                f"Alunos sem avaliacao - {dados['escola']} - {dados['turma']} - "
                f"{dados['bimestre']}Ã‚Âº bimestre"
            ]
        )
        planilha.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cabecalho))
        planilha.append(cabecalho)

        for aluno in dados["alunos"]:
            planilha.append(
                [
                    aluno["numero_chamada"],
                    aluno["aluno"],
                    "Nao realizou"
                    if aluno["ausente_dia_1"]
                    else "-" if aluno["ausente_dia_1"] is None else "Realizou",
                    "Nao realizou"
                    if aluno["ausente_dia_2"]
                    else "-" if aluno["ausente_dia_2"] is None else "Realizou",
                    aluno["status"],
                ]
            )

        titulo = planilha[1][0]
        titulo.font = Font(bold=True, size=14)
        titulo.alignment = Alignment(horizontal="center")

        for celula in planilha[2]:
            celula.font = Font(bold=True)
            celula.fill = preenchimento_cabecalho
            celula.alignment = Alignment(horizontal="center")

        for coluna in planilha.columns:
            largura = max(len(str(celula.value or "")) for celula in coluna) + 2
            planilha.column_dimensions[get_column_letter(coluna[0].column)].width = min(
                max(largura, 12),
                42,
            )

        for linha in planilha.iter_rows(min_row=3):
            for celula in linha:
                celula.alignment = Alignment(vertical="center")

    arquivo = BytesIO()
    workbook.save(arquivo)
    arquivo.seek(0)
    return arquivo


@app.get("/")
def home():
    return {"mensagem": "Backend do Sistema de Gabarito funcionando"}


@app.get("/teste-banco")
def teste_banco(db: Session = Depends(get_db)):
    resultado = db.execute(text("select nome from escolas")).fetchall()
    escolas = [linha[0] for linha in resultado]

    return {"conectado": True, "escolas": escolas}


@app.get("/escolas")
def listar_escolas(db: Session = Depends(get_db)):
    return db.query(models.Escola).order_by(models.Escola.nome).all()


@app.get("/turmas/{escola_id}")
def listar_turmas(escola_id: str, db: Session = Depends(get_db)):
    return (
        db.query(models.Turma)
        .filter(models.Turma.escola_id == escola_id)
        .order_by(models.Turma.nome)
        .all()
    )


@app.get("/alunos/{turma_id}")
def listar_alunos(turma_id: str, db: Session = Depends(get_db)):
    return (
        db.query(models.Aluno)
        .filter(models.Aluno.turma_id == turma_id)
        .order_by(models.Aluno.numero_chamada)
        .all()
    )


@app.get("/modelo-prova")
def buscar_modelo_prova(
    escola_id: str,
    bimestre: int,
    dia: int,
    db: Session = Depends(get_db),
):
    modelo = _buscar_modelo_prova(db, escola_id, bimestre, dia)

    disciplinas = (
        db.query(models.DisciplinaProva)
        .filter(models.DisciplinaProva.modelo_prova_id == modelo.id)
        .order_by(models.DisciplinaProva.ordem)
        .all()
    )

    return {"modelo": modelo, "disciplinas": disciplinas}


@app.post("/gabarito")
def salvar_gabarito(
    escola_id: str = Form(...),
    bimestre: int = Form(...),
    dia: int = Form(...),
    serie: int = Form(...),
    codigo_gabarito: str = Form(GABARITO_PADRAO),
    respostas: str = Form(...),
    db: Session = Depends(get_db),
):
    modelo = _buscar_modelo_prova(db, escola_id, bimestre, dia)
    codigo_gabarito = _normalizar_codigo_gabarito(codigo_gabarito)

    disciplinas = (
        db.query(models.DisciplinaProva)
        .filter(models.DisciplinaProva.modelo_prova_id == modelo.id)
        .order_by(models.DisciplinaProva.ordem)
        .all()
    )

    if not disciplinas:
        raise HTTPException(status_code=404, detail="Disciplinas da prova nao encontradas")

    respostas_lista = [_normalizar_resposta(resposta) for resposta in respostas.split(",") if resposta.strip()]

    if any(resposta is None for resposta in respostas_lista):
        raise HTTPException(status_code=400, detail="Use somente alternativas A, B, C, D ou X")

    total_esperado = sum(disciplina.quantidade_questoes for disciplina in disciplinas)
    if len(respostas_lista) != total_esperado:
        raise HTTPException(
            status_code=400,
            detail=f"Quantidade de respostas invalida. Esperado: {total_esperado}. Recebido: {len(respostas_lista)}.",
        )

    try:
        (
            db.query(models.Gabarito)
            .filter(models.Gabarito.modelo_prova_id == modelo.id)
            .filter(models.Gabarito.serie == serie)
            .filter(models.Gabarito.codigo_gabarito == codigo_gabarito)
            .delete(synchronize_session=False)
        )

        numero_questao = 1
        indice_resposta = 0

        for disciplina in disciplinas:
            for _ in range(disciplina.quantidade_questoes):
                gabarito = models.Gabarito(
                    modelo_prova_id=modelo.id,
                    serie=serie,
                    codigo_gabarito=codigo_gabarito,
                    numero_questao=numero_questao,
                    disciplina=disciplina.disciplina,
                    resposta_correta=respostas_lista[indice_resposta],
                )
                db.add(gabarito)

                numero_questao += 1
                indice_resposta += 1

        db.flush()
        gabaritos_salvos = _buscar_gabarito(db, modelo.id, serie, codigo_gabarito)
        resultados_recalculados = _recalcular_resultados_por_gabarito(
            db,
            modelo,
            serie,
            codigo_gabarito,
            gabaritos_salvos,
        )
        db.commit()
    except SQLAlchemyError as exc:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Erro ao salvar gabarito: {exc}") from exc

    return {
        "mensagem": "Gabarito salvo com sucesso",
        "serie": serie,
        "codigo_gabarito": codigo_gabarito,
        "total_questoes": total_esperado,
        "resultados_recalculados": resultados_recalculados,
    }


@app.get("/gabarito")
def buscar_gabarito(
    escola_id: str,
    bimestre: int,
    dia: int,
    serie: int,
    codigo_gabarito: str = GABARITO_PADRAO,
    db: Session = Depends(get_db),
):
    modelo = _buscar_modelo_prova(db, escola_id, bimestre, dia)
    return _buscar_gabarito(db, modelo.id, serie, codigo_gabarito)


@app.post("/corrigir-manual")
def corrigir_manual(
    aluno_id: str = Form(...),
    escola_id: str = Form(...),
    bimestre: int = Form(...),
    dia: int = Form(...),
    respostas: str = Form(...),
    codigo_gabarito: str | None = Form(None),
    db: Session = Depends(get_db),
):
    modelo = _buscar_modelo_prova(db, escola_id, bimestre, dia)
    serie, codigo_gabarito = _buscar_contexto_correcao(db, aluno_id, modelo, codigo_gabarito)
    gabaritos = _buscar_gabarito(db, modelo.id, serie, codigo_gabarito)
    respostas_lista = [_normalizar_resposta(resposta) for resposta in respostas.split(",") if resposta.strip()]

    if any(resposta is None for resposta in respostas_lista):
        raise HTTPException(status_code=400, detail="Use somente alternativas A, B, C ou D")

    if len(respostas_lista) < len(gabaritos) or (
        len(respostas_lista) > len(gabaritos) and not _modelo_eh_daniela(db, modelo)
    ):
        raise HTTPException(
            status_code=400,
            detail=f"Quantidade de respostas invalida. Esperado: {len(gabaritos)}. Recebido: {len(respostas_lista)}.",
        )

    respostas_lista = respostas_lista[: len(gabaritos)]

    respostas_detectadas = {
        indice + 1: resposta for indice, resposta in enumerate(respostas_lista)
    }

    try:
        acertos, respostas_salvas = _comparar_e_salvar_respostas(
            db,
            aluno_id,
            modelo.id,
            gabaritos,
            respostas_detectadas,
        )
        total_questoes = len(gabaritos)
        nota_dia = _salvar_resultado_aluno(
            db,
            aluno_id,
            modelo,
            serie,
            codigo_gabarito,
            acertos,
            total_questoes,
        )
        resumo_global = _calcular_nota_global_bimestre(db, aluno_id, escola_id, bimestre)
        db.commit()
    except SQLAlchemyError as exc:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Erro ao salvar correcao: {exc}") from exc

    return {
        "mensagem": "Prova corrigida com sucesso",
        "aluno_id": aluno_id,
        "escola_id": escola_id,
        "bimestre": bimestre,
        "dia": dia,
        "serie": serie,
        "codigo_gabarito": codigo_gabarito,
        "acertos": acertos,
        "total_questoes": total_questoes,
        "nota_dia": nota_dia,
        **resumo_global,
        "respostas_salvas": respostas_salvas,
        "gabarito_lido_tabulado": _tabular_respostas(respostas_salvas),
    }


@app.post("/corrigir-foto")
async def corrigir_foto(
    aluno_id: str = Form(...),
    escola_id: str = Form(...),
    bimestre: int = Form(...),
    dia: int = Form(...),
    codigo_gabarito: str | None = Form(None),
    foto: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    modelo = _buscar_modelo_prova(db, escola_id, bimestre, dia)
    serie, codigo_gabarito = _buscar_contexto_correcao(db, aluno_id, modelo, codigo_gabarito)
    gabaritos = _buscar_gabarito(db, modelo.id, serie, codigo_gabarito)
    total_questoes = len(gabaritos)

    nome_arquivo, caminho_arquivo = _salvar_upload(foto, await foto.read())

    resultado_processamento, erro = processar_folha(caminho_arquivo, total_questoes=total_questoes)
    if erro:
        raise HTTPException(
            status_code=422,
            detail={
                "mensagem": erro,
                "imagem_original": nome_arquivo,
                "modelo_prova_id": str(modelo.id),
            },
        )

    caminho_folha = os.path.join(UPLOAD_DIR, f"folha_{nome_arquivo}")
    caminho_threshold = os.path.join(UPLOAD_DIR, f"threshold_{nome_arquivo}")

    cv2.imwrite(caminho_folha, resultado_processamento["folha"])
    cv2.imwrite(caminho_threshold, resultado_processamento["folha_threshold"])

    try:
        respostas_detectadas, debug_respostas = ler_respostas_grade_fixa(
            resultado_processamento["folha_threshold"],
            total_questoes=total_questoes,
        )
    except Exception as exc:
        raise HTTPException(status_code=422, detail=f"Erro ao ler respostas: {exc}") from exc

    try:
        acertos, respostas_salvas = _comparar_e_salvar_respostas(
            db,
            aluno_id,
            modelo.id,
            gabaritos,
            respostas_detectadas,
        )
        nota_dia = _salvar_resultado_aluno(
            db,
            aluno_id,
            modelo,
            serie,
            codigo_gabarito,
            acertos,
            total_questoes,
        )
        resumo_global = _calcular_nota_global_bimestre(db, aluno_id, escola_id, bimestre)
        db.commit()
    except SQLAlchemyError as exc:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Erro ao salvar respostas: {exc}") from exc

    return {
        "mensagem": "Prova corrigida automaticamente",
        "aluno_id": aluno_id,
        "escola_id": escola_id,
        "bimestre": bimestre,
        "dia": dia,
        "serie": serie,
        "codigo_gabarito": codigo_gabarito,
        "imagem_original": nome_arquivo,
        "imagem_alinhada": f"folha_{nome_arquivo}",
        "imagem_threshold": f"threshold_{nome_arquivo}",
        "modelo_prova_id": str(modelo.id),
        "respostas_detectadas": respostas_detectadas,
        "debug_respostas": debug_respostas,
        "respostas_salvas": respostas_salvas,
        "gabarito_lido_tabulado": _tabular_respostas(respostas_salvas),
        "acertos": acertos,
        "total_questoes": total_questoes,
        "nota_dia": nota_dia,
        **resumo_global,
    }


@app.post("/corrigir-foto-existente")
def corrigir_foto_existente(
    aluno_id: str = Form(...),
    escola_id: str = Form(...),
    bimestre: int = Form(...),
    dia: int = Form(...),
    nome_arquivo: str = Form(...),
    codigo_gabarito: str | None = Form(None),
    db: Session = Depends(get_db),
):
    modelo = _buscar_modelo_prova(db, escola_id, bimestre, dia)
    serie, codigo_gabarito = _buscar_contexto_correcao(db, aluno_id, modelo, codigo_gabarito)
    gabaritos = _buscar_gabarito(db, modelo.id, serie, codigo_gabarito)
    total_questoes = len(gabaritos)

    nome_arquivo = os.path.basename(nome_arquivo)

    if nome_arquivo.startswith(("folha_", "threshold_", "debug_", "cinza_", "bordas_")):
        raise HTTPException(
            status_code=400,
            detail="Informe o nome da foto original, nao uma imagem gerada pelo processamento.",
        )

    caminho_arquivo = os.path.join(UPLOAD_DIR, nome_arquivo)

    if not os.path.exists(caminho_arquivo):
        raise HTTPException(status_code=404, detail="Imagem nao encontrada em uploads")

    resultado_processamento, erro = processar_folha(caminho_arquivo, total_questoes=total_questoes)
    if erro:
        raise HTTPException(
            status_code=422,
            detail={
                "mensagem": erro,
                "imagem_original": nome_arquivo,
                "modelo_prova_id": str(modelo.id),
            },
        )

    caminho_folha = os.path.join(UPLOAD_DIR, f"folha_{nome_arquivo}")
    caminho_threshold = os.path.join(UPLOAD_DIR, f"threshold_{nome_arquivo}")

    cv2.imwrite(caminho_folha, resultado_processamento["folha"])
    cv2.imwrite(caminho_threshold, resultado_processamento["folha_threshold"])

    try:
        respostas_detectadas, debug_respostas = ler_respostas_grade_fixa(
            resultado_processamento["folha_threshold"],
            total_questoes=total_questoes,
        )
    except Exception as exc:
        raise HTTPException(status_code=422, detail=f"Erro ao ler respostas: {exc}") from exc

    try:
        acertos, respostas_salvas = _comparar_e_salvar_respostas(
            db,
            aluno_id,
            modelo.id,
            gabaritos,
            respostas_detectadas,
        )
        nota_dia = _salvar_resultado_aluno(
            db,
            aluno_id,
            modelo,
            serie,
            codigo_gabarito,
            acertos,
            total_questoes,
        )
        resumo_global = _calcular_nota_global_bimestre(db, aluno_id, escola_id, bimestre)
        db.commit()
    except SQLAlchemyError as exc:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Erro ao salvar respostas: {exc}") from exc

    return {
        "mensagem": "Imagem existente relida e prova corrigida automaticamente",
        "aluno_id": aluno_id,
        "escola_id": escola_id,
        "bimestre": bimestre,
        "dia": dia,
        "serie": serie,
        "codigo_gabarito": codigo_gabarito,
        "imagem_original": nome_arquivo,
        "imagem_alinhada": f"folha_{nome_arquivo}",
        "imagem_threshold": f"threshold_{nome_arquivo}",
        "modelo_prova_id": str(modelo.id),
        "respostas_detectadas": respostas_detectadas,
        "debug_respostas": debug_respostas,
        "respostas_salvas": respostas_salvas,
        "gabarito_lido_tabulado": _tabular_respostas(respostas_salvas),
        "acertos": acertos,
        "total_questoes": total_questoes,
        "nota_dia": nota_dia,
        **resumo_global,
    }


@app.get("/respostas-aluno")
def buscar_respostas_aluno(
    aluno_id: str,
    escola_id: str,
    bimestre: int,
    dia: int,
    db: Session = Depends(get_db),
):
    modelo = _buscar_modelo_prova(db, escola_id, bimestre, dia)
    resultado = (
        db.query(models.ResultadoAluno)
        .filter(models.ResultadoAluno.aluno_id == aluno_id)
        .filter(models.ResultadoAluno.modelo_prova_id == modelo.id)
        .first()
    )

    respostas = (
        db.query(models.RespostaAluno)
        .filter(models.RespostaAluno.aluno_id == aluno_id)
        .filter(models.RespostaAluno.modelo_prova_id == modelo.id)
        .order_by(models.RespostaAluno.numero_questao)
        .all()
    )

    respostas_salvas = [_resposta_aluno_para_dict(resposta) for resposta in respostas]
    total_questoes = resultado.total_questoes if resultado else len(respostas_salvas)
    acertos = resultado.acertos if resultado else sum(1 for resposta in respostas_salvas if resposta["acertou"])
    nota_dia = resultado.nota_global if resultado else _calcular_nota(acertos, total_questoes)
    resumo_global = _calcular_nota_global_bimestre(db, aluno_id, escola_id, bimestre)

    return {
        "aluno_id": aluno_id,
        "escola_id": escola_id,
        "bimestre": bimestre,
        "dia": dia,
        "modelo_prova_id": str(modelo.id),
        "codigo_gabarito": resultado.codigo_gabarito if resultado else None,
        "respostas_salvas": respostas_salvas,
        "gabarito_lido_tabulado": _tabular_respostas(respostas_salvas),
        "acertos": acertos,
        "total_questoes": total_questoes,
        "nota_dia": nota_dia,
        **resumo_global,
    }


@app.delete("/correcao-aluno")
def excluir_correcao_aluno(
    aluno_id: str,
    escola_id: str,
    bimestre: int,
    dia: int,
    db: Session = Depends(get_db),
):
    modelo = _buscar_modelo_prova(db, escola_id, bimestre, dia)

    respostas_removidas = (
        db.query(models.RespostaAluno)
        .filter(models.RespostaAluno.aluno_id == aluno_id)
        .filter(models.RespostaAluno.modelo_prova_id == modelo.id)
        .delete(synchronize_session=False)
    )
    resultados_removidos = (
        db.query(models.ResultadoAluno)
        .filter(models.ResultadoAluno.aluno_id == aluno_id)
        .filter(models.ResultadoAluno.modelo_prova_id == modelo.id)
        .delete(synchronize_session=False)
    )

    if not respostas_removidas and not resultados_removidos:
        raise HTTPException(status_code=404, detail="Correcao do aluno nao encontrada")

    db.commit()
    resumo_global = _calcular_nota_global_bimestre(db, aluno_id, escola_id, bimestre)

    return {
        "mensagem": "Correcao excluida com sucesso",
        "aluno_id": aluno_id,
        "escola_id": escola_id,
        "bimestre": bimestre,
        "dia": dia,
        "respostas_removidas": respostas_removidas,
        "resultados_removidos": resultados_removidos,
        **resumo_global,
    }


@app.get("/resultados-alunos")
def listar_resultados_alunos(
    turma_id: str,
    escola_id: str,
    bimestre: int,
    dia: int | None = None,
    db: Session = Depends(get_db),
):
    modelos_query = (
        db.query(models.ModeloProva)
        .filter(models.ModeloProva.escola_id == escola_id)
        .filter(models.ModeloProva.bimestre == bimestre)
    )

    if dia is not None:
        modelos_query = modelos_query.filter(models.ModeloProva.dia == dia)

    modelos = modelos_query.order_by(models.ModeloProva.dia).all()

    if not modelos:
        raise HTTPException(status_code=404, detail="Modelo de prova nao encontrado")

    modelo_ids = [modelo.id for modelo in modelos]
    dias_por_modelo = {modelo.id: modelo.dia for modelo in modelos}

    alunos = (
        db.query(models.Aluno)
        .filter(models.Aluno.turma_id == turma_id)
        .order_by(models.Aluno.numero_chamada)
        .all()
    )
    aluno_ids = [aluno.id for aluno in alunos]

    if not aluno_ids:
        return []

    resultados = (
        db.query(models.ResultadoAluno)
        .filter(models.ResultadoAluno.modelo_prova_id.in_(modelo_ids))
        .filter(models.ResultadoAluno.aluno_id.in_(aluno_ids))
        .all()
    )
    resultados_bimestre = (
        db.query(models.ResultadoAluno)
        .filter(models.ResultadoAluno.escola_id == escola_id)
        .filter(models.ResultadoAluno.bimestre == bimestre)
        .filter(models.ResultadoAluno.aluno_id.in_(aluno_ids))
        .all()
    )
    respostas = (
        db.query(models.RespostaAluno)
        .filter(models.RespostaAluno.modelo_prova_id.in_(modelo_ids))
        .filter(models.RespostaAluno.aluno_id.in_(aluno_ids))
        .order_by(models.RespostaAluno.numero_questao)
        .all()
    )

    respostas_por_aluno = {}
    for resposta in respostas:
        respostas_por_aluno.setdefault(str(resposta.aluno_id), []).append(
            _resposta_aluno_para_dict(resposta)
        )

    disciplinas_modelo = (
        db.query(models.DisciplinaProva)
        .filter(models.DisciplinaProva.modelo_prova_id.in_(modelo_ids))
        .order_by(models.DisciplinaProva.ordem)
        .all()
        if modelo_ids
        else []
    )
    nomes_disciplinas = []
    disciplinas_por_modelo = {}
    for disciplina in disciplinas_modelo:
        disciplinas_por_modelo.setdefault(disciplina.modelo_prova_id, []).append(
            disciplina.disciplina
        )
        if disciplina.disciplina not in nomes_disciplinas:
            nomes_disciplinas.append(disciplina.disciplina)

    def calcular_media_disciplinas_respostas(respostas_aluno):
        return _calcular_media_notas_disciplinas(
            {
                disciplina: resumo["nota"]
                for disciplina, resumo in calcular_notas_disciplinas_respostas(
                    respostas_aluno
                ).items()
            }
        )

    def calcular_notas_disciplinas_respostas(respostas_aluno):
        resumo_disciplinas = {
            disciplina: {"acertos": 0, "total": 0}
            for disciplina in nomes_disciplinas
        }
        for resposta in respostas_aluno:
            disciplina = resposta["disciplina"] or "Sem disciplina"
            if disciplina not in resumo_disciplinas:
                resumo_disciplinas[disciplina] = {"acertos": 0, "total": 0}
            resumo_disciplinas[disciplina]["total"] += 1
            if resposta["acertou"]:
                resumo_disciplinas[disciplina]["acertos"] += 1

        notas_disciplinas = {
            disciplina: {
                **resumo,
                "nota": _calcular_nota(resumo["acertos"], resumo["total"])
                if resumo["total"]
                else 0,
            }
            for disciplina, resumo in resumo_disciplinas.items()
        }
        return notas_disciplinas

    linhas_por_aluno = {}
    globais_por_aluno = {}
    dias_por_aluno = {}

    for resultado in resultados_bimestre:
        aluno_id = str(resultado.aluno_id)
        if aluno_id not in globais_por_aluno:
            globais_por_aluno[aluno_id] = {
                "acertos_global": 0,
                "total_questoes_global": 0,
            }

        globais_por_aluno[aluno_id]["acertos_global"] += resultado.acertos
        globais_por_aluno[aluno_id]["total_questoes_global"] += resultado.total_questoes

        dia_resultado = resultado.dia
        dias_por_aluno.setdefault(aluno_id, {})[dia_resultado] = {
            "acertos": resultado.acertos,
            "total_questoes": resultado.total_questoes,
            "nota": resultado.nota_global,
            "codigo_gabarito": resultado.codigo_gabarito,
        }

    for aluno_id, resumo in globais_por_aluno.items():
        resumo["nota_global"] = calcular_media_disciplinas_respostas(
            respostas_por_aluno.get(aluno_id, [])
        )

    disciplinas_por_aluno = {}
    for aluno_id, respostas_salvas in respostas_por_aluno.items():
        disciplinas_por_aluno[aluno_id] = calcular_notas_disciplinas_respostas(respostas_salvas)

    for resultado in resultados_bimestre:
        if resultado.total_questoes != 0:
            continue

        aluno_id = str(resultado.aluno_id)
        disciplinas_aluno = disciplinas_por_aluno.setdefault(
            aluno_id,
            {
                disciplina: {"acertos": 0, "total": 0, "nota": 0}
                for disciplina in nomes_disciplinas
            },
        )

        for disciplina in disciplinas_por_modelo.get(resultado.modelo_prova_id, []):
            disciplinas_aluno[disciplina] = {
                "acertos": 0,
                "total": 0,
                "nota": resultado.nota_global,
            }

        globais_por_aluno.setdefault(
            aluno_id,
            {
                "acertos_global": 0,
                "total_questoes_global": 0,
            },
        )
        globais_por_aluno[aluno_id]["nota_global"] = _calcular_media_notas_disciplinas(
            {
                disciplina: resumo["nota"]
                for disciplina, resumo in disciplinas_aluno.items()
            }
        )

    for resultado in resultados:
        aluno_id = str(resultado.aluno_id)
        resumo_global = globais_por_aluno.get(
            aluno_id,
            {
                "acertos_global": resultado.acertos,
                "total_questoes_global": resultado.total_questoes,
                "nota_global": calcular_media_disciplinas_respostas(
                    respostas_por_aluno.get(aluno_id, [])
                ),
            },
        )
        acertos_linha = resultado.acertos if dia is not None else resumo_global["acertos_global"]
        total_linha = resultado.total_questoes if dia is not None else resumo_global["total_questoes_global"]
        linhas_por_aluno[aluno_id] = {
            "aluno_id": aluno_id,
            "modelo_prova_id": str(resultado.modelo_prova_id),
            "escola_id": str(resultado.escola_id),
            "bimestre": resultado.bimestre,
            "dia": resultado.dia,
            "serie": resultado.serie,
            "codigo_gabarito": resultado.codigo_gabarito,
            "acertos": acertos_linha,
            "total_questoes": total_linha,
            "nota_dia": resultado.nota_global if dia is not None else None,
            **resumo_global,
            "disciplinas": disciplinas_por_aluno.get(aluno_id, {}),
            "dias_modelo": sorted({modelo.dia for modelo in modelos}),
            "resultados_dias": dias_por_aluno.get(aluno_id, {}),
            "respostas_salvas": respostas_por_aluno.get(aluno_id, []),
        }

    for aluno_id, respostas_salvas in respostas_por_aluno.items():
        if aluno_id in linhas_por_aluno:
            continue

        total_questoes = len(respostas_salvas)
        acertos = sum(1 for resposta in respostas_salvas if resposta["acertou"])
        resumo_global = globais_por_aluno.get(
            aluno_id,
            {
                "acertos_global": acertos,
                "total_questoes_global": total_questoes,
                "nota_global": calcular_media_disciplinas_respostas(respostas_salvas),
            },
        )
        linhas_por_aluno[aluno_id] = {
            "aluno_id": aluno_id,
            "modelo_prova_id": ",".join(str(modelo_id) for modelo_id in modelo_ids),
            "escola_id": escola_id,
            "bimestre": bimestre,
            "dia": dia,
            "serie": None,
            "acertos": acertos,
            "total_questoes": total_questoes,
            "nota_dia": _calcular_nota(acertos, total_questoes),
            **resumo_global,
            "disciplinas": disciplinas_por_aluno.get(aluno_id, {}),
            "dias_modelo": sorted({modelo.dia for modelo in modelos}),
            "resultados_dias": dias_por_aluno.get(aluno_id, {}),
            "respostas_salvas": respostas_salvas,
        }

    return [
        linhas_por_aluno[str(aluno.id)]
        for aluno in alunos
        if str(aluno.id) in linhas_por_aluno
    ]


@app.get("/relatorio-ausentes-turma")
def relatorio_ausentes_turma(
    turma_id: str,
    escola_id: str,
    bimestre: int,
    db: Session = Depends(get_db),
):
    return _montar_relatorio_ausentes_turma(db, turma_id, escola_id, bimestre)


@app.get("/relatorio-ausentes-turma-xlsx")
def baixar_relatorio_ausentes_turma_xlsx(
    turma_id: str,
    escola_id: str,
    bimestre: int,
    db: Session = Depends(get_db),
):
    dados = _montar_relatorio_ausentes_turma(db, turma_id, escola_id, bimestre)
    arquivo = _montar_excel_relatorio_ausentes(dados)
    nome_arquivo = f"ausentes_{dados['turma'].replace(' ', '_')}_{bimestre}_bimestre.xlsx"

    return StreamingResponse(
        arquivo,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nome_arquivo}"'},
    )


@app.get("/relatorio-ausentes-escola-xlsx")
def baixar_relatorio_ausentes_escola_xlsx(
    escola_id: str,
    bimestre: int,
    db: Session = Depends(get_db),
):
    arquivo = _montar_excel_relatorios_ausentes_escola(db, escola_id, bimestre)
    nome_arquivo = f"ausentes_todas_turmas_{bimestre}_bimestre.xlsx"

    return StreamingResponse(
        arquivo,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nome_arquivo}"'},
    )


@app.get("/resultado-final-excel")
def baixar_resultado_final_excel(
    escola_id: str,
    bimestre: int,
    db: Session = Depends(get_db),
):
    escola = db.query(models.Escola).filter(models.Escola.id == escola_id).first()
    if not escola:
        raise HTTPException(status_code=404, detail="Escola nao encontrada")

    return baixar_resultado_final_xlsx(escola_id, bimestre, db)

    disciplinas, linhas = _montar_resultado_final_escola(db, escola_id, bimestre)
    arquivo = _montar_excel_resultado_final(escola.nome, bimestre, disciplinas, linhas)
    arquivo.write(";".join(["Turma", "Nº", "Aluno", *disciplinas, "Media Global", "Status"]))

    for linha in linhas:
        valores = [
            linha["turma"],
            str(linha["numero_chamada"]),
            linha["aluno"],
            *[_formatar_nota_csv(linha["disciplinas"].get(disciplina)) for disciplina in disciplinas],
            _formatar_nota_csv(linha["nota_global"]),
            linha["status"],
        ]
        arquivo.write(";".join(_celula_csv(valor) for valor in valores))
        arquivo.write("\n")

    arquivo.seek(0)
    nome_arquivo = f"resultado_final_{bimestre}_bimestre.csv"

    return StreamingResponse(
        iter([arquivo.getvalue()]),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{nome_arquivo}"'},
    )


@app.get("/resultado-final-xlsx")
def baixar_resultado_final_xlsx(
    escola_id: str,
    bimestre: int,
    db: Session = Depends(get_db),
):
    escola = db.query(models.Escola).filter(models.Escola.id == escola_id).first()
    if not escola:
        raise HTTPException(status_code=404, detail="Escola nao encontrada")

    disciplinas, linhas = _montar_resultado_final_escola(db, escola_id, bimestre)
    arquivo = _montar_excel_resultado_final(escola.nome, bimestre, disciplinas, linhas)
    nome_arquivo = f"resultado_final_{bimestre}_bimestre.xlsx"

    return StreamingResponse(
        arquivo,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nome_arquivo}"'},
    )


@app.patch("/nota-aluno")
def editar_nota_aluno(
    aluno_id: str = Form(...),
    escola_id: str = Form(...),
    bimestre: int = Form(...),
    dia: int = Form(...),
    nota: float = Form(...),
    db: Session = Depends(get_db),
):
    if nota < 0 or nota > 10:
        raise HTTPException(status_code=400, detail="A nota deve estar entre 0 e 10")

    modelo = _buscar_modelo_prova(db, escola_id, bimestre, dia)
    serie = _buscar_serie_aluno(db, aluno_id)
    resultado = (
        db.query(models.ResultadoAluno)
        .filter(models.ResultadoAluno.aluno_id == aluno_id)
        .filter(models.ResultadoAluno.modelo_prova_id == modelo.id)
        .first()
    )

    if not resultado:
        resultado = models.ResultadoAluno(
            aluno_id=aluno_id,
            modelo_prova_id=modelo.id,
            escola_id=modelo.escola_id,
            bimestre=modelo.bimestre,
            dia=modelo.dia,
            serie=serie,
            codigo_gabarito=GABARITO_PADRAO,
            acertos=0,
            total_questoes=0,
            nota_global=round(nota, 1),
        )
        db.add(resultado)
    else:
        resultado.escola_id = modelo.escola_id
        resultado.bimestre = modelo.bimestre
        resultado.dia = modelo.dia
        resultado.serie = serie
        resultado.nota_global = round(nota, 1)

    db.commit()
    resumo_global = _calcular_nota_global_bimestre(db, aluno_id, escola_id, bimestre)

    return {
        "mensagem": "Nota atualizada com sucesso",
        "aluno_id": aluno_id,
        "escola_id": escola_id,
        "bimestre": bimestre,
        "dia": dia,
        "nota_dia": resultado.nota_global,
        **resumo_global,
    }


@app.patch("/acertos-aluno")
def editar_acertos_aluno(
    aluno_id: str = Form(...),
    escola_id: str = Form(...),
    bimestre: int = Form(...),
    dia: int = Form(...),
    acertos: int = Form(...),
    db: Session = Depends(get_db),
):
    modelo = _buscar_modelo_prova(db, escola_id, bimestre, dia)
    resultado = (
        db.query(models.ResultadoAluno)
        .filter(models.ResultadoAluno.aluno_id == aluno_id)
        .filter(models.ResultadoAluno.modelo_prova_id == modelo.id)
        .first()
    )

    if not resultado:
        raise HTTPException(status_code=404, detail="Resultado do aluno nao encontrado")

    if acertos < 0 or acertos > resultado.total_questoes:
        raise HTTPException(
            status_code=400,
            detail=f"Acertos devem ficar entre 0 e {resultado.total_questoes}",
        )

    respostas = (
        db.query(models.RespostaAluno)
        .filter(models.RespostaAluno.aluno_id == aluno_id)
        .filter(models.RespostaAluno.modelo_prova_id == modelo.id)
        .order_by(models.RespostaAluno.numero_questao)
        .all()
    )

    for indice, resposta in enumerate(respostas):
        resposta.acertou = indice < acertos

    resultado.acertos = acertos
    resultado.nota_global = _calcular_nota(acertos, resultado.total_questoes)
    db.commit()
    resumo_global = _calcular_nota_global_bimestre(db, aluno_id, escola_id, bimestre)

    return {
        "mensagem": "Acertos atualizados com sucesso",
        "aluno_id": aluno_id,
        "escola_id": escola_id,
        "bimestre": bimestre,
        "dia": dia,
        "acertos": resultado.acertos,
        "total_questoes": resultado.total_questoes,
        "nota_dia": resultado.nota_global,
        **resumo_global,
    }


@app.patch("/nota-adaptada")
def salvar_nota_adaptada(
    aluno_id: str = Form(...),
    escola_id: str = Form(...),
    bimestre: int = Form(...),
    dia: int = Form(...),
    acertos_disciplinas: str = Form(...),
    db: Session = Depends(get_db),
):
    modelo = _buscar_modelo_prova(db, escola_id, bimestre, dia)
    serie = _buscar_serie_aluno(db, aluno_id)

    disciplinas = (
        db.query(models.DisciplinaProva)
        .filter(models.DisciplinaProva.modelo_prova_id == modelo.id)
        .order_by(models.DisciplinaProva.ordem)
        .all()
    )

    if not disciplinas:
        raise HTTPException(status_code=404, detail="Disciplinas da prova nao encontradas")

    try:
        acertos_por_disciplina = json.loads(acertos_disciplinas)
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=400, detail="Acertos por disciplina invalidos") from exc

    total_acertos = 0
    total_questoes = 0

    try:
        (
            db.query(models.RespostaAluno)
            .filter(models.RespostaAluno.aluno_id == aluno_id)
            .filter(models.RespostaAluno.modelo_prova_id == modelo.id)
            .delete(synchronize_session=False)
        )

        numero_questao = 1
        for disciplina in disciplinas:
            acertos = acertos_por_disciplina.get(disciplina.disciplina)
            if acertos is None:
                acertos = acertos_por_disciplina.get(disciplina.sigla)

            if acertos is None:
                raise HTTPException(
                    status_code=400,
                    detail=f"Informe os acertos de {disciplina.disciplina}",
                )

            try:
                acertos = int(acertos)
            except (TypeError, ValueError) as exc:
                raise HTTPException(
                    status_code=400,
                    detail=f"Acertos de {disciplina.disciplina} devem ser um numero inteiro",
                ) from exc

            if acertos < 0 or acertos > disciplina.quantidade_questoes:
                raise HTTPException(
                    status_code=400,
                    detail=(
                        f"Acertos de {disciplina.disciplina} devem ficar entre 0 "
                        f"e {disciplina.quantidade_questoes}"
                    ),
                )

            total_acertos += acertos
            total_questoes += disciplina.quantidade_questoes

            for indice in range(disciplina.quantidade_questoes):
                db.add(
                    models.RespostaAluno(
                        aluno_id=aluno_id,
                        modelo_prova_id=modelo.id,
                        numero_questao=numero_questao,
                        disciplina=disciplina.disciplina,
                        resposta_aluno=None,
                        resposta_correta=None,
                        acertou=indice < acertos,
                    )
                )
                numero_questao += 1

        nota_dia = _salvar_resultado_aluno(
            db,
            aluno_id,
            modelo,
            serie,
            GABARITO_ADAPTADA,
            total_acertos,
            total_questoes,
        )
        resumo_global = _calcular_nota_global_bimestre(db, aluno_id, escola_id, bimestre)
        db.commit()
    except SQLAlchemyError as exc:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Erro ao salvar prova adaptada: {exc}") from exc

    respostas = (
        db.query(models.RespostaAluno)
        .filter(models.RespostaAluno.aluno_id == aluno_id)
        .filter(models.RespostaAluno.modelo_prova_id == modelo.id)
        .order_by(models.RespostaAluno.numero_questao)
        .all()
    )
    respostas_salvas = [_resposta_aluno_para_dict(resposta) for resposta in respostas]

    return {
        "mensagem": "Prova adaptada salva com sucesso",
        "aluno_id": aluno_id,
        "escola_id": escola_id,
        "bimestre": bimestre,
        "dia": dia,
        "serie": serie,
        "codigo_gabarito": GABARITO_ADAPTADA,
        "modelo_prova_id": str(modelo.id),
        "respostas_salvas": respostas_salvas,
        "gabarito_lido_tabulado": _tabular_respostas(respostas_salvas),
        "acertos": total_acertos,
        "total_questoes": total_questoes,
        "nota_dia": nota_dia,
        **resumo_global,
    }


@app.get("/resultados")
def resultados(
    escola: str | None = None,
    turma: str | None = None,
    db: Session = Depends(get_db),
):
    sql = """
        select *
        from vw_resultado_final
        where (:escola is null or escola = :escola)
        and (:turma is null or turma = :turma)
        order by turma, numero_chamada
    """

    resultado = db.execute(text(sql), {"escola": escola, "turma": turma}).mappings().all()
    return list(resultado)
